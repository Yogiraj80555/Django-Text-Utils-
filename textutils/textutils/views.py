from django.http import HttpResponse,HttpResponseRedirect
from django.shortcuts import render
from django.views.generic import TemplateView
from django.core.files.storage import FileSystemStorage

def errorFun(request, exception):
	return HttpResponse("<h1>Service Unavailable(503)</h1>")

def index(request):
	"""with open('text.txt') as f:
		data = f.read()
		return HttpResponse(data);"""
	return render(request,'index.html')
	#return HttpResponse("data");

def about(request):
	return HttpResponse("<h1>This is about Page</h1>");


def delete(request):
	get = request.POST.get('text','get default value if not Text');
	punc = request.POST.get('remove','off');
	upr = request.POST.get('uppercase','off');
	newl = request.POST.get('newline','off');
	rmvspc = request.POST.get('removespace','off');

	punch = ".,?!'\":;-@#$%^&*_()[]}{";
	data = process(get,punc,upr,newl,rmvspc);
	print(data)
	dct = {"purpose":punc,"analyzed_text":data};
	return render(request,'analyzed.html',dct);



def process(string,punc,upper,newline,extraspace):
	if punc == "on":
		strng=""
		for i in string:
			if i not in ".,?!'\":;-@#$%^&*_()[]}{":
				strng+=i
		string = strng
	if upper == "on":
		string = string.upper();
	if newline == "on":
		strs=""
		print(string)
		for i in string:
			if i != "\n" and i !="\r":
				strs+=i
		string = strs
		print(string)
	if extraspace == "on":
		strs = ""
		for i,val in enumerate(string):
			if not(string[i] == " " and string[i+1] == " "):
				strs += val
		string = strs

	return string;
	
	
#####################  Excel to json ###############################	
	
def excel(request):
	navclass = "<nav class='navbar navbar-expand-lg navbar-light' style='background-color: #cfe9fc;'>"
	welcomemsg = "background-color: #cef1fd"
	data = {'jsondata':False,'error':False,'title':"XLSX to JSON",'appName':"Excel to JSON",'navClasses':str(navclass),'welcme-color':str(welcomemsg)}
	if request.method == 'POST':
		try:
			upload_file = request.FILES['uploadfile']
			fs = FileSystemStorage()
			fs.save(upload_file.name,upload_file)
			data["jsondata"] = parsingdata(upload_file.name,request.POST.get('editfile','no'))
			removefile()
		except Exception as e:
			print("Exception is: ",e);
			data['error'] = e;
	print(data)
	return render(request,'exceltojson.html',data)

def parsingdata(name,sheet):
	from openpyxl import Workbook
	from openpyxl import load_workbook
	##
	##
	##
	path = "/home/yogiraj80555/TextUtil/media/" + name
	#print("path is:",path)
	workbook = load_workbook(filename=path);
	#print(workbook.sheetnames)
	
	if sheet not in workbook.sheetnames:
		return "Invalid Sheet Name -> "+ sheet;
	workbook.active =  workbook.sheetnames.index(sheet)
	sheet = workbook.active
	products=""
	for col in sheet.iter_rows(min_row=1,max_row=1,values_only=True): #getting columns
		products=col;
	main_Dct = {}
	for row in sheet.iter_rows(min_row=2,min_col=1,values_only=True):
		dct = {}
		for i in range(1,len(products)):
			dct[products[i]] = str(row[i]);
		main_Dct[row[0]] = dct;
	import json
	return json.dumps(main_Dct)
	
def removefile():
	import os
	import glob
	##
	#here can be chage at production
	path = "/home/yogiraj80555/TextUtil/media"
	files = glob.glob(os.path.join(path+'/*.*'))
	for i in files:
		os.remove(i)
	print(files)
	
####################################################################


#######################  Anagram  ##################################
f = ""
def anagram(request):
	try:
		navclass = "<nav class='navbar navbar-expand-lg  navbar-light' style='background-color: #cef1fd;'>"
		welcomeclr = "'background-color: #e6f8fe'"
		data = {'jsondata':False,'error':False,'title':"Anagram",'appName':"Anagram",'navClasses':str(navclass),'welcmeclr':str(welcomeclr)}
		import os
		path = "/home/yogiraj80555/TextUtil/media/data/words.txt"
		global f
		print("Length of Global F ",len(f))
		if len(f) < 2:
			file =  open(os.path.join(path),'r')
			f = file.read();
			file.close()
		data['words'] = f.split()
		try:
			if request.method == 'POST':
				data['jsondata'] = anagramLogic(f.split(),request.POST.get('editfield','0'))
				print(data['jsondata'])
		except Exception as e: 
			data['error'] = e
	except Exception as e:
		print("My Error is ->",e)
	return render(request,"anagram.html",data) 
	
	
def anagramLogic(words,word):
	if word is '0' or len(word) < 4:
		raise Exception("Please enter valid word or at least 4 letter word ")
		 
	word = word.lower()
	for i in word:
		words = [s for s in words if i in s and len(s) < (len(word)+1) ]   
	print("Condition of data ",words,"->",len(words))
	if len(words)>1:
		if word in words:
			words.remove(word)
		data = " We found ",len(words)," Anagrams of  '"+word+"' is follow: <br>"+str(set(words))
		return str(data)
	else:
		return "We not found any anagram"
		
		
###################################################################

#######################  Infix to prepostfix  #####################
def infix(request):
	navclass = "<nav class='navbar navbar-expand-lg navbar-dark ' style='background-color: #034fa0;'>"
	welcomeclr = "'background-color: #0078d4; color: #fff;'"
	data = {'title':"Infix To Pre-Postfix conversion",'appName':"Infix Conversion",'navClasses':str(navclass),'welcmeclr':str(welcomeclr)}
	if request.method == "POST":
		try:
			expression = request.POST.get('infiexpression' , '0');
			conversion = request.POST.get('exampleRadios', '0');
			data['answers'] = doConversion(expression.strip(),conversion);
			print("\n\nExpression: ",data['answers'],"\n",expression,"\n")
			
			
		except Exception as e:
			print("Infix conversion error ",e);
			data['error'] = e;
	
	
	return render(request,"infixtoprepostfix.html",data)
	
	
def doConversion(infix,conversion):
	#validation of expression
	import re
	pattern=re.compile("^[a-zA-Z\-+*/()^]+$")
	if pattern.fullmatch(infix) is not None:
		if conversion == "prefix":
			from .logics.infixtoprepostfix import infixtoprefix as prefix
			return prefix.getInfixString(infix)
		elif conversion == "postfix":
			from .logics.infixtoprepostfix import infixtopostfix as postfix 
			return postfix.getInfixString(infix)
		else: return "Invalid Conversion String";
	return "Invalid Infix Expression";


	
	
	
###################################################################
